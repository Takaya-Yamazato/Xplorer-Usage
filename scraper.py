"""
IEEE Xplore Usage Statistics Scraper
=====================================
対象ジャーナル:
  - IEICE Transactions on Communications (punumber=10400553)
  - IEICE Communications Express      (punumber=10250155)

各論文のUsage（Abstract Views + Full Text Views + PDF Downloads など）を
年・月ごとに集計してExcelファイルに出力します。

依存パッケージのインストール:
  python3 -m venv venv
  source venv/bin/activate
  pip install playwright pandas openpyxl tqdm matplotlib japanize-matplotlib setuptools
  playwright install chromium

実行方法:
  python scraper.py

オプション:
  --headed         ブラウザを表示して実行（デフォルト: 非表示）
  --output FILE    出力Excelファイル名（デフォルト: ieice_usage.xlsx）
  --delay FLOAT    リクエスト間の待機秒数（デフォルト: 2.0）
  --max-pages INT  各ジャーナルの最大ページ数（0=全ページ、デフォルト: 0）
  --doi DOI        特定のDOIのUsageを取得する（複数指定可）
    python scraper.py --doi 10.23919/comex.2023XBL0092 10.23919/comex.2023XBL0098
  --pub-year INT   特定の出版年の論文のみを取得する（例: 2024）
    python scraper.py --pub-year 2024
  --usage-year INT 月別Usageシートで特定の利用年のみを出力する（例: 2024）
    python scraper.py --usage-year 2024


"""

import argparse
import json
import time
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import pandas as pd
from tqdm import tqdm

# ---------------------------------------------------------------------------
# 設定
# ---------------------------------------------------------------------------
JOURNALS = {
    "IEICE Trans. Commun.": "10400553",
    "IEICE Commun. Express": "10250155",
}

BASE_URL = "https://ieeexplore.ieee.org"

# IEEE Xplore REST API エンドポイント
SEARCH_API = BASE_URL + "/rest/search"
ARTICLE_METRICS_API = BASE_URL + "/rest/document/{article_id}/metrics"
ARTICLE_REFERENCES_API = BASE_URL + "/rest/document/{article_id}/references"
ARTICLE_AUTHORS_API = BASE_URL + "/rest/document/{article_id}/authors"

ROWS_PER_PAGE = 100  # 最大100件/ページ　IEEE Xploreの検索APIが 「1ページあたりの最大取得件数を100件までに制限している」


# ---------------------------------------------------------------------------
# メイン処理
# ---------------------------------------------------------------------------
def parse_args():
    p = argparse.ArgumentParser(description="IEEE Xplore IEICE Usage Scraper")
    p.add_argument("--headed", action="store_true", help="ブラウザを表示して実行（デフォルト: 非表示）")
    p.add_argument("--output", default="ieice_usage.xlsx", help="出力ファイル名")
    p.add_argument("--delay", type=float, default=2.0, help="リクエスト間待機秒数")
    p.add_argument("--max-pages", type=int, default=0, help="最大ページ数(0=全件)")
    p.add_argument("--doi", nargs="+", metavar="DOI", help="特定のDOIのみをスクレイピングする（複数指定可）")
    p.add_argument("--pub-year", type=int, default=0, help="特定の出版年の論文のみを取得する（例: 2024）")
    p.add_argument("--usage-year", type=int, default=0, help="特定の利用年（Usage Year）の月別データのみを出力する（例: 2024）")
    return p.parse_args()


def fetch_json_via_playwright(page, url: str, params: dict = None) -> dict:
    """Playwrightのページオブジェクトを使ってJSON APIを叩く。"""
    method = "POST" if params else "GET"
    max_retries = 3
    for attempt in range(max_retries):
        try:
            result = page.evaluate(
                """async ([url, method, params]) => {
                    try {
                        const options = {
                            method: method,
                            headers: {
                                'Accept': 'application/json, text/plain, */*',
                                'X-Requested-With': 'XMLHttpRequest',
                            }
                        };
                        if (method === 'POST' && params) {
                            options.headers['Content-Type'] = 'application/json';
                            options.body = JSON.stringify(params);
                        }
                        const resp = await fetch(url, options);
                        if (!resp.ok) return {error: resp.status};
                        return await resp.json();
                    } catch (e) {
                        return {error: "fetch_failed", message: e.toString()};
                    }
                }""",
                [url, method, params],
            )
            if isinstance(result, dict) and result.get("error") == "fetch_failed":
                raise Exception(result.get("message", "Fetch failed"))
            return result or {}
        except Exception as e:
            if attempt < max_retries - 1:
                time.sleep(2 * (attempt + 1))  # 失敗したら 2秒, 4秒 と待機してリトライ
            else:
                print(f"\n  [ERROR] API request failed after {max_retries} retries: {url} - {e}")
                return {"error": str(e)}


def get_all_articles(page, punumber: str, max_pages: int, delay: float, pub_year: int = 0) -> list[dict]:
    """指定ジャーナルの全論文リストを取得する。"""
    articles = []
    page_num = 1

    # まず1ページ目を取得して総件数を確認
    data = fetch_json_via_playwright(page, SEARCH_API, {
        "newsearch": "true",
        "queryText": "",
        "punumber": punumber,
        "pageNumber": page_num,
        "rowsPerPage": ROWS_PER_PAGE,
        "sortType": "newest",
    })

    if "error" in data:
        print(f"  [ERROR] API応答エラー: {data}")
        return []

    total = data.get("totalRecords", 0)
    total_pages = (total + ROWS_PER_PAGE - 1) // ROWS_PER_PAGE
    if max_pages > 0:
        total_pages = min(total_pages, max_pages)

    if not pub_year:
        print(f"  総論文数: {total} 件 / {total_pages} ページ")

    def extract_articles(data):
        return data.get("records", [])

    fetched = extract_articles(data)
    articles.extend(fetched)

    # pub_year 指定時の早期終了判定（新しい順で取得しているため、すべて指定年より古くなったら終了）
    def should_stop(fetched_arts):
        if not pub_year: return False
        for art in fetched_arts:
            y, _ = parse_publication_date(art)
            if y and y >= pub_year:
                return False
        return True

    if total_pages > 1 and not should_stop(fetched):
        for page_num in tqdm(range(2, total_pages + 1), desc="  論文リスト取得"):
            time.sleep(delay)
            data = fetch_json_via_playwright(page, SEARCH_API, {
                "newsearch": "false",
                "queryText": "",
                "punumber": punumber,
                "pageNumber": page_num,
                "rowsPerPage": ROWS_PER_PAGE,
                "sortType": "newest",
            })
            fetched = extract_articles(data)
            articles.extend(fetched)
            if should_stop(fetched):
                break

    # 重複排除（APIのページまたぎ等で同じ論文が複数含まれるのを防ぐため）
    seen = set()
    unique_articles = []
    for art in articles:
        aid = str(art.get("articleNumber", art.get("arnumber", "")))
        y, _ = parse_publication_date(art)
        if pub_year and y != pub_year:
            continue
        if aid and aid not in seen:
            seen.add(aid)
            unique_articles.append(art)
            
    if pub_year:
        print(f"  総論文数（{pub_year}年）: {len(unique_articles)} 件")

    return unique_articles


def get_article_metrics(page, article_id: str) -> dict:
    """個別論文のUsageメトリクスを取得する。"""
    url = ARTICLE_METRICS_API.format(article_id=article_id)
    return fetch_json_via_playwright(page, url)

def get_article_references(page, article_id: str) -> dict:
    """個別論文の参考文献(References)データを取得する。"""
    url = ARTICLE_REFERENCES_API.format(article_id=article_id)
    return fetch_json_via_playwright(page, url)

def get_article_authors(page, article_id: str) -> dict:
    """個別論文の著者データを取得する。"""
    url = ARTICLE_AUTHORS_API.format(article_id=article_id)
    return fetch_json_via_playwright(page, url)


def parse_publication_date(article: dict) -> tuple[int | None, int | None]:
    """論文の出版年・月を抽出する。"""
    # publicationDate フィールド例: "2024-03", "January 2024", "2024" など
    pub_date = article.get("publicationDate", "")
    pub_year = article.get("publicationYear")
    pub_month = None

    # "YYYY-MM" 形式
    m = re.match(r"(\d{4})-(\d{2})", pub_date)
    if m:
        return int(m.group(1)), int(m.group(2))

    # "Month YYYY" 形式
    month_map = {
        "jan": 1, "feb": 2, "mar": 3, "apr": 4,
        "may": 5, "jun": 6, "jul": 7, "aug": 8,
        "sep": 9, "oct": 10, "nov": 11, "dec": 12,
    }
    m = re.match(r"([A-Za-z]+)\s+(\d{4})", pub_date)
    if m:
        month_name = m.group(1).lower()[:3]
        pub_month = month_map.get(month_name)
        return int(m.group(2)), pub_month

    # issueDate などを試みる
    issue_date = article.get("issueDate", "")
    m = re.match(r"(\d{4})-(\d{2})", issue_date)
    if m:
        return int(m.group(1)), int(m.group(2))

    # 年だけ
    if pub_year:
        return int(pub_year), None

    return None, None


def collect_usage(page, journal_name: str, punumber: str,
                  max_pages: int, delay: float, pub_year: int = 0) -> list[dict]:
    """1ジャーナルの全Usage情報を収集してレコードのリストを返す。"""
    print(f"\n{'='*60}")
    print(f"ジャーナル: {journal_name}  (punumber={punumber})")
    print(f"{'='*60}")

    articles = get_all_articles(page, punumber, max_pages, delay, pub_year)

    records = []
    errors = 0

    for art in tqdm(articles, desc="  Usage取得"):
        article_id = str(art.get("articleNumber", art.get("arnumber", "")))
        if not article_id:
            continue

        year, month = parse_publication_date(art)
        if pub_year and year != pub_year:
            continue

        title = art.get("articleTitle", art.get("documentTitle", art.get("title", "")))
        doi = art.get("doi", "")
        citation_count = art.get("citationCount", 0)

        # Usage情報はメトリクスAPIから取得
        time.sleep(delay)
        metrics = get_article_metrics(page, article_id)

        # 参考文献数情報を取得
        time.sleep(delay)
        refs_data = get_article_references(page, article_id)
        reference_count = 0
        ieice_ref_count = 0
        if isinstance(refs_data, dict):
            references = refs_data.get("references", [])
            reference_count = len(references)
            # 参考文献内のIEICE文献数をカウント
            for ref in references:
                pub_title = ref.get("publicationTitle", "")
                ref_text = ref.get("text", "")
                if "ieice" in pub_title.lower() or "ieice" in ref_text.lower():
                    ieice_ref_count += 1

        ieice_ref_ratio = ieice_ref_count / reference_count if reference_count > 0 else 0.0

        # 著者情報を取得
        time.sleep(delay)
        authors_data = get_article_authors(page, article_id)
        authors_list = authors_data.get("authors", art.get("authors", [])) if isinstance(authors_data, dict) else art.get("authors", [])
        first_author_name = ""
        first_author_affil = ""
        if authors_list:
            first_a = authors_list[0]
            first_author_name = first_a.get("name", first_a.get("preferredName", ""))
            affil = first_a.get("affiliation") or first_a.get("affiliations") or first_a.get("authorAffiliations") or ""
            if isinstance(affil, list):
                affil = ", ".join(str(x) for x in affil)
            first_author_affil = affil

        if "error" in metrics or not metrics:
            errors += 1
            # フォールバック: article自体に usageCount や downloadCount フィールドがある場合
            usage_count = art.get("usageCount", art.get("downloadCount", art.get("citationCount", 0)))
            records.append({
                "journal": journal_name,
                "article_id": article_id,
                "doi": doi,
                "title": title,
                "first_author": first_author_name,
                "first_author_affiliation": first_author_affil,
                "year": year,
                "month": month,
                "total_usage": usage_count,
                "citation_count": citation_count,
                "reference_count": reference_count,
                "ieice_ref_count": ieice_ref_count,
                "ieice_ref_ratio": ieice_ref_ratio,
                "biblio": [],
            })
            continue

        # メトリクスから各種Usage値を取得
        # IEEE Xploreのメトリクス構造に合わせて抽出
        inner_metrics = metrics.get("metrics", {})
        biblio = inner_metrics.get("biblio", [])
        total_usage = (
            metrics.get("usageCount")
            or metrics.get("totalUsageCount")
            or inner_metrics.get("totalDownloads")
            or art.get("usageCount")
            or art.get("downloadCount")
            or 0
        )

        records.append({
            "journal": journal_name,
            "article_id": article_id,
            "doi": doi,
            "title": title,
            "first_author": first_author_name,
            "first_author_affiliation": first_author_affil,
            "year": year,
            "month": month,
            "total_usage": total_usage,
            "citation_count": citation_count,
            "reference_count": reference_count,
            "ieice_ref_count": ieice_ref_count,
            "ieice_ref_ratio": ieice_ref_ratio,
            "biblio": biblio,
        })

    print(f"  完了。エラー件数: {errors}")
    return records


def collect_usage_by_doi(page, dois: list[str], delay: float) -> list[dict]:
    """指定されたDOIのUsage情報を収集してレコードのリストを返す。"""
    print(f"\n{'='*60}")
    print(f"DOI指定モード: {len(dois)}件")
    print(f"{'='*60}")

    records = []
    errors = 0

    for i, doi in enumerate(dois):
        doi = doi.strip()
        print(f"  [{i+1}/{len(dois)}] DOI検索中: {doi}")
        time.sleep(delay)

        # DOIで検索
        search_result = fetch_json_via_playwright(page, SEARCH_API, {
            "newsearch": "true",
            "queryText": f"doi:{doi}",
            "pageNumber": 1,
            "rowsPerPage": 5,
        })

        found = search_result.get("records", [])
        matched = [
            a for a in found
            if a.get("doi", "").lower().strip() == doi.lower().strip()
        ]
        if not matched and found:
            matched = [found[0]]
            print(f"    [WARN] 完全一致なし。先頭の結果を使用: {found[0].get('doi', '')}")

        if not matched:
            print(f"    [ERROR] 文献が見つかりませんでした: {doi}")
            errors += 1
            continue

        art = matched[0]
        article_id = str(art.get("articleNumber", art.get("arnumber", "")))
        year, month = parse_publication_date(art)
        title = art.get("articleTitle", art.get("documentTitle", art.get("title", "")))
        journal_name = art.get("publicationTitle", "Unknown Journal")
        citation_count = art.get("citationCount", 0)

        print(f"    論文ID: {article_id}")

        # メトリクス取得
        time.sleep(delay)
        metrics = get_article_metrics(page, article_id)

        # 参考文献数情報を取得
        time.sleep(delay)
        refs_data = get_article_references(page, article_id)
        reference_count = 0
        ieice_ref_count = 0
        if isinstance(refs_data, dict):
            references = refs_data.get("references", [])
            reference_count = len(references)
            # 参考文献内のIEICE文献数をカウント
            for ref in references:
                pub_title = ref.get("publicationTitle", "")
                ref_text = ref.get("text", "")
                if "ieice" in pub_title.lower() or "ieice" in ref_text.lower():
                    ieice_ref_count += 1

        ieice_ref_ratio = ieice_ref_count / reference_count if reference_count > 0 else 0.0

        # 著者情報を取得
        time.sleep(delay)
        authors_data = get_article_authors(page, article_id)
        authors_list = authors_data.get("authors", art.get("authors", [])) if isinstance(authors_data, dict) else art.get("authors", [])
        first_author_name = ""
        first_author_affil = ""
        if authors_list:
            first_a = authors_list[0]
            first_author_name = first_a.get("name", first_a.get("preferredName", ""))
            affil = first_a.get("affiliation") or first_a.get("affiliations") or first_a.get("authorAffiliations") or ""
            if isinstance(affil, list):
                affil = ", ".join(str(x) for x in affil)
            first_author_affil = affil

        if "error" in metrics or not metrics:
            errors += 1
            usage_count = art.get("usageCount", art.get("downloadCount", art.get("citationCount", 0)))
            records.append({
                "journal": journal_name, "article_id": article_id,
                "doi": doi, "title": title,
                "first_author": first_author_name,
                "first_author_affiliation": first_author_affil,
                "year": year, "month": month,
                "total_usage": usage_count,
                "citation_count": citation_count,
                "reference_count": reference_count,
                "ieice_ref_count": ieice_ref_count,
                "ieice_ref_ratio": ieice_ref_ratio,
                "biblio": [],
            })
            print(f"    Usage -> Total: {usage_count} (フォールバック)")
            continue

        inner_metrics = metrics.get("metrics", {})
        biblio = inner_metrics.get("biblio", [])
        total_usage = metrics.get("usageCount") or metrics.get("totalUsageCount") or inner_metrics.get("totalDownloads") or art.get("usageCount") or art.get("downloadCount") or 0

        records.append({
            "journal": journal_name,
            "article_id": article_id,
            "doi": doi,
            "title": title,
            "first_author": first_author_name,
            "first_author_affiliation": first_author_affil,
            "year": year,
            "month": month,
            "total_usage": total_usage,
            "citation_count": citation_count,
            "reference_count": reference_count,
            "ieice_ref_count": ieice_ref_count,
            "ieice_ref_ratio": ieice_ref_ratio,
            "biblio": biblio,
        })
        print(f"    Usage -> Total={total_usage}")

    print(f"  完了。エラー件数: {errors}")
    return records


def build_pivot_table(df: pd.DataFrame, value_col, journal_name: str, agg_type: str = "sum") -> pd.DataFrame:
    """年・月ごとのピボットテーブルを作成する。"""
    df_j = df[df["journal"] == journal_name].copy()
    df_j = df_j.dropna(subset=["year"])
    df_j["year"] = df_j["year"].astype(int)
    df_j["month"] = df_j["month"].fillna(0).astype(int)

    if df_j.empty:
        return pd.DataFrame()

    if agg_type == "mean":
        pivot_sum = df_j.pivot_table(index="year", columns="month", values=value_col, aggfunc="sum", fill_value=0)
        pivot_count = df_j.pivot_table(index="year", columns="month", values=value_col, aggfunc="count", fill_value=0)

        sum_row = pivot_sum.sum(axis=1)
        count_row = pivot_count.sum(axis=1)

        sum_col = pivot_sum.sum(axis=0)
        count_col = pivot_count.sum(axis=0)

        total_sum = sum_row.sum()
        total_count = count_row.sum()

        pivot = (pivot_sum / pivot_count).fillna(0).round(1)
        pivot["合計"] = (sum_row / count_row).fillna(0).round(1)

        total_series = (sum_col / count_col).fillna(0).round(1)
        total_series["合計"] = round(total_sum / total_count, 1) if total_count > 0 else 0
        pivot.loc["合計"] = total_series
    elif agg_type == "ratio":
        num_col, den_col = value_col
        pivot_num = df_j.pivot_table(index="year", columns="month", values=num_col, aggfunc="sum", fill_value=0)
        pivot_den = df_j.pivot_table(index="year", columns="month", values=den_col, aggfunc="sum", fill_value=0)

        pivot_num, pivot_den = pivot_num.align(pivot_den, fill_value=0)

        sum_num_row = pivot_num.sum(axis=1)
        sum_den_row = pivot_den.sum(axis=1)
        sum_num_col = pivot_num.sum(axis=0)
        sum_den_col = pivot_den.sum(axis=0)
        total_num = sum_num_row.sum()
        total_den = sum_den_row.sum()

        pivot = pivot_num.divide(pivot_den).replace([float('inf'), -float('inf')], 0).fillna(0).round(1)
        pivot["合計"] = (sum_num_row / sum_den_row).replace([float('inf'), -float('inf')], 0).fillna(0).round(1)

        total_series = (sum_num_col / sum_den_col).replace([float('inf'), -float('inf')], 0).fillna(0).round(1)
        total_series["合計"] = round(total_num / total_den, 1) if total_den > 0 else 0
        pivot.loc["合計"] = total_series
    else:
        pivot = df_j.pivot_table(
            index="year",
            columns="month",
            values=value_col,
            aggfunc=agg_type,
            fill_value=0,
        )
        pivot["合計"] = pivot.sum(axis=1)
        pivot.loc["合計"] = pivot.sum()

    # 月のカラム名を整形
    month_names = {
        0: "月不明",
        1: "1月", 2: "2月", 3: "3月", 4: "4月",
        5: "5月", 6: "6月", 7: "7月", 8: "8月",
        9: "9月", 10: "10月", 11: "11月", 12: "12月",
        "合計": "合計"
    }
    pivot.columns = [month_names.get(c, str(c)) for c in pivot.columns]
    pivot.index.name = "年"

    return pivot


def save_to_excel(df_all: pd.DataFrame, records: list[dict], output_path: str, usage_year: int = 0):
    """結果をExcelファイルに保存する。"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    writer = pd.ExcelWriter(output_path, engine="openpyxl")

    # ---- Sheet 1: 集計サマリー（年・月ごと、ジャーナル別） ----
    row_cursor = 1
    ws_name = "Usage集計"

    metric_configs = [
        ("total_usage", "Number of usage per issue", "sum"),
        ("total_usage", "Number of articles", "count"),
        ("total_usage", "Average usage per article", "mean"),
        ("citation_count", "Number of citations per issue", "sum"),
        ("citation_count", "Average citation per article", "mean"),
        (("total_usage", "citation_count"), "Average usage per citation", "ratio"),
        ("ieice_ref_count", "Number of references to IEICE journals", "sum"),
    ]

    unique_journals = df_all["journal"].unique()
    for journal_name in unique_journals:
        col_cursor = 0
        max_rows = 0
        for metric_col, metric_label, agg_type in metric_configs:
            key = f"{journal_name} / {metric_label}"
            pivot = build_pivot_table(df_all, metric_col, journal_name, agg_type)
            if not pivot.empty:
                pivot.to_excel(writer, sheet_name=ws_name, startrow=row_cursor, startcol=col_cursor)
                ws = writer.sheets[ws_name]
                ws.cell(row=row_cursor, column=col_cursor+1, value=f"【{key}】")
                ws.cell(row=row_cursor, column=col_cursor+1).font = Font(bold=True, size=11)
                max_rows = max(max_rows, len(pivot) + 4)
                col_cursor += len(pivot.columns) + 2
        if max_rows > 0:
            row_cursor += max_rows

    # ---- グラフ作成用データの追加 ----
    df_chart = df_all.copy()
    df_chart["year"] = pd.to_numeric(df_chart["year"], errors='coerce')
    df_chart["month"] = pd.to_numeric(df_chart["month"], errors='coerce')
    df_chart = df_chart.dropna(subset=["year", "month"])
    df_chart = df_chart[df_chart["month"] > 0]

    if not df_chart.empty:
        # yyyy/mm の形式で文字列を作成
        df_chart["date_str"] = df_chart["year"].astype(int).astype(str) + "/" + df_chart["month"].astype(int).astype(str).str.zfill(2)
        
        # 昇順ソートした全年月を取得
        unique_dates = sorted(df_chart["date_str"].unique())
        
        row_cursor += 3  # 空白行をあける
        
        if ws_name in writer.sheets:
            ws = writer.sheets[ws_name]
            
            ws.cell(row=row_cursor, column=1, value="【グラフ作成用データ】")
            ws.cell(row=row_cursor, column=1).font = Font(bold=True, size=11)
            row_cursor += 1
            
            target_metrics = [
                ("Number of usage per issue", "total_usage", "sum"),
                ("Average usage per article", "total_usage", "mean"),
                ("Number of citations per issue", "citation_count", "sum"),
                ("Number of articles", "total_usage", "count"),
                ("Average reference count", "reference_count", "mean"),
                ("Average ieice_ref_count per issue", "ieice_ref_count", "mean"),
                ("Average ieice_ref_ratio per issue", "ieice_ref_ratio", "mean"),
            ]
            
            for journal_name in unique_journals:
                df_j = df_chart[df_chart["journal"] == journal_name]
                ws.cell(row=row_cursor, column=1, value=journal_name)
                ws.cell(row=row_cursor, column=1).font = Font(bold=True)
                row_cursor += 1
                
                # ヘッダー行 (年月)
                for j, d_str in enumerate(unique_dates):
                    ws.cell(row=row_cursor, column=j+2, value=d_str)
                row_cursor += 1
                
                # 月ごとに集計
                if not df_j.empty:
                    monthly_stats = df_j.groupby('date_str').agg(
                        usage_sum=('total_usage', 'sum'),
                        usage_count=('total_usage', 'count'),
                        usage_mean=('total_usage', 'mean'),
                        citation_sum=('citation_count', 'sum'),
                        reference_mean=('reference_count', 'mean'),
                        ieice_ref_mean=('ieice_ref_count', 'mean'),
                        ieice_ref_ratio_mean=('ieice_ref_ratio', 'mean')
                    )
                else:
                    monthly_stats = pd.DataFrame()

                for label, col, agg in target_metrics:
                    ws.cell(row=row_cursor, column=1, value=label)
                    
                    if agg == "sum" and col == "total_usage":
                        stat_col = "usage_sum"
                    elif agg == "count" and col == "total_usage":
                        stat_col = "usage_count"
                    elif agg == "mean" and col == "total_usage":
                        stat_col = "usage_mean"
                    elif agg == "sum" and col == "citation_count":
                        stat_col = "citation_sum"
                    elif agg == "mean" and col == "reference_count":
                        stat_col = "reference_mean"
                    elif agg == "mean" and col == "ieice_ref_count":
                        stat_col = "ieice_ref_mean"
                    elif agg == "mean" and col == "ieice_ref_ratio":
                        stat_col = "ieice_ref_ratio_mean"
                    else:
                        stat_col = None

                    for j, d_str in enumerate(unique_dates):
                        val = 0
                        if stat_col and d_str in monthly_stats.index:
                            val = monthly_stats.loc[d_str, stat_col]
                            if pd.isna(val):
                                val = 0
                            elif agg == "mean":
                                val = round(val, 1)
                        ws.cell(row=row_cursor, column=j+2, value=val)
                    row_cursor += 1
                row_cursor += 1  # ジャーナル間に1行空ける

    # ---- Sheet 2: 論文ごとの詳細 ----
    df_main = df_all.drop(columns=["biblio"], errors="ignore")
    df_main.to_excel(writer, sheet_name="論文詳細", index=False)

    # ---- Sheet 3: 年別合計サマリー（2ジャーナル横並び） ----
    yearly_rows = []
    for journal_name in unique_journals:
        df_j = df_all[df_all["journal"] == journal_name].dropna(subset=["year"])
        if df_j.empty:
            continue
        yearly = df_j.groupby("year")[
            ["total_usage", "citation_count", "reference_count", "ieice_ref_count"]
        ].sum()
        yearly.columns = [f"{journal_name} / {c}" for c in yearly.columns]
        yearly_rows.append(yearly)

    if yearly_rows:
        yearly_summary = pd.concat(yearly_rows, axis=1).fillna(0).astype(int)
        yearly_summary.index.name = "年"
        yearly_summary.loc["合計"] = yearly_summary.sum()
        yearly_summary.to_excel(writer, sheet_name="年別合計")

    # ---- Sheet 4: 月別Usage詳細 ----
    biblio_rows = []
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    for r in records:
        for b in r.get("biblio", []):
            if usage_year and str(b.get("year")) != str(usage_year):
                continue
            row = {
                "journal": r.get("journal", ""),
                "article_id": r.get("article_id", ""),
                "doi": r.get("doi", ""),
                "title": r.get("title", ""),
                "usage_year": b.get("year", ""),
            }
            for m in months:
                val = b.get(m, "-")
                try:
                    row[m] = int(val) if val != "-" else 0
                except ValueError:
                    row[m] = 0
            row["yearly_total"] = b.get("yearToDateDownloads", 0)
            biblio_rows.append(row)
            
    if biblio_rows:
        df_biblio = pd.DataFrame(biblio_rows)
        df_biblio.to_excel(writer, sheet_name="月別Usage", index=False)

    writer.close()
    print(f"\n✅ Excelファイルを保存しました: {output_path}")


def save_charts(df_all: pd.DataFrame, output_path: str):
    """月別のUsage推移グラフを生成してファイルに保存する。"""
    try:
        import matplotlib.pyplot as plt
        import matplotlib.dates as mdates
        import japanize_matplotlib
    except ImportError as e:
        print("\n" + "="*60)
        print("🚨 [WARN] グラフ描画ライブラリがインストールされていません。")
        print(f"詳細なエラー: {e}")
        print("PNG画像（グラフ）を出力するには、以下のコマンドを実行してください:")
        print("  pip install matplotlib japanize-matplotlib setuptools")
        print("="*60 + "\n")
        return

    unique_journals = df_all["journal"].unique()
    for journal_name in unique_journals:
        df_j = df_all[df_all["journal"] == journal_name].copy()
        if df_j.empty:
            continue

        # より安全に日付型へ変換し、NaN行を削除
        df_j["year"] = pd.to_numeric(df_j["year"], errors='coerce')
        df_j["month"] = pd.to_numeric(df_j["month"], errors='coerce')
        df_j = df_j.dropna(subset=["year", "month"])
        df_j = df_j[df_j["month"] > 0]

        if df_j.empty:
            print(f"  [WARN] {journal_name} のグラフをスキップしました（月が特定できる論文データがありません）")
            continue
        df_j["date"] = pd.to_datetime(df_j["year"].astype(int).astype(str) + '-' + df_j["month"].astype(int).astype(str).str.zfill(2) + '-01')

        # 月ごとに集計
        monthly_stats = df_j.groupby('date').agg(
            usage_sum=('total_usage', 'sum'),
            usage_mean=('total_usage', 'mean'),
            usage_count=('total_usage', 'count'),
            citation_sum=('citation_count', 'sum'),
            citation_mean=('citation_count', 'mean')
        ).reset_index()
        monthly_stats = monthly_stats.sort_values('date')
        monthly_stats['usage_per_citation'] = (monthly_stats['usage_sum'] / monthly_stats['citation_sum']).replace([float('inf'), -float('inf')], 0).fillna(0)

        if monthly_stats.empty:
            continue

        base_name = Path(output_path).stem
        journal_suffix = journal_name.replace(' ', '_').replace('.', '')

        # 1. Usage グラフ
        fig, ax1 = plt.subplots(figsize=(18, 8))
        ax1.bar(monthly_stats['date'], monthly_stats['usage_sum'], width=20, label='Number of usage per issue', color='skyblue')
        ax1.set_xlabel('年月')
        ax1.set_ylabel('Number of usage per issue (合計ダウンロード数)')
        ax2 = ax1.twinx()
        ax2.plot(monthly_stats['date'], monthly_stats['usage_mean'], color='orangered', marker='o', linestyle='-', label='Average usage per article')
        ax2.set_ylabel('Average usage per article (論文あたり平均)')
        fig.suptitle(f'月別Usage推移 - {journal_name}', fontsize=16)
        ax1.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m'))
        fig.autofmt_xdate(rotation=45)
        lines, labels = ax1.get_legend_handles_labels()
        lines2, labels2 = ax2.get_legend_handles_labels()
        ax2.legend(lines + lines2, labels + labels2, loc='upper left')
        fig.tight_layout(rect=[0, 0.03, 1, 0.95])
        chart_filename_usage = f"{base_name}_{journal_suffix}_usage.png"
        plt.savefig(chart_filename_usage)
        print(f"✅ グラフを保存しました: {chart_filename_usage}")
        plt.close(fig)

        # 2. Citation グラフ
        fig, ax1 = plt.subplots(figsize=(18, 8))
        ax1.bar(monthly_stats['date'], monthly_stats['citation_sum'], width=20, label='Number of citations per issue', color='lightgreen')
        ax1.set_xlabel('年月')
        ax1.set_ylabel('Number of citations per issue (合計被引用数)')
        ax2 = ax1.twinx()
        ax2.plot(monthly_stats['date'], monthly_stats['citation_mean'], color='green', marker='o', linestyle='-', label='Average citation per article')
        ax2.set_ylabel('Average citation per article (論文あたり平均)')
        fig.suptitle(f'月別Citation推移 - {journal_name}', fontsize=16)
        ax1.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m'))
        fig.autofmt_xdate(rotation=45)
        lines, labels = ax1.get_legend_handles_labels()
        lines2, labels2 = ax2.get_legend_handles_labels()
        ax2.legend(lines + lines2, labels + labels2, loc='upper left')
        fig.tight_layout(rect=[0, 0.03, 1, 0.95])
        chart_filename_citation = f"{base_name}_{journal_suffix}_citation.png"
        plt.savefig(chart_filename_citation)
        print(f"✅ グラフを保存しました: {chart_filename_citation}")
        plt.close(fig)

        # 3. Usage per Citation グラフ
        fig, ax1 = plt.subplots(figsize=(18, 8))
        ax1.plot(monthly_stats['date'], monthly_stats['usage_per_citation'], color='purple', marker='s', linestyle='-', label='Average usage per citation')
        ax1.set_xlabel('年月')
        ax1.set_ylabel('Average usage per citation (被引用1件あたりのUsage)')
        fig.suptitle(f'月別 Average usage per citation - {journal_name}', fontsize=16)
        ax1.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m'))
        fig.autofmt_xdate(rotation=45)
        ax1.legend(loc='upper left')
        fig.tight_layout(rect=[0, 0.03, 1, 0.95])
        chart_filename_upc = f"{base_name}_{journal_suffix}_usage_per_citation.png"
        plt.savefig(chart_filename_upc)
        print(f"✅ グラフを保存しました: {chart_filename_upc}")
        plt.close(fig)


def main():
    args = parse_args()

    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print("Playwrightがインストールされていません。")
        print("  pip install playwright")
        print("  playwright install chromium")
        sys.exit(1)

    print("=" * 60)
    print("IEEE Xplore IEICE Usage Scraper")
    print("=" * 60)
    print(f"出力ファイル: {args.output}")
    print(f"待機時間: {args.delay} 秒")
    print(f"最大ページ数: {'全件' if args.max_pages == 0 else args.max_pages}")
    print()

    all_records = []

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=not args.headed)
        context = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1280, "height": 900},
        )
        page = context.new_page()

        # IEEE Xplore トップページを開いてCookieを取得
        print("IEEE Xplore を開いています...")
        page.goto(BASE_URL, wait_until="networkidle", timeout=30000)
        time.sleep(2)

        if args.doi:
            records = collect_usage_by_doi(page, args.doi, args.delay)
            all_records.extend(records)
        else:
            for journal_name, punumber in JOURNALS.items():
                records = collect_usage(
                    page, journal_name, punumber, args.max_pages, args.delay, args.pub_year
                )
                all_records.extend(records)

        browser.close()

    if not all_records:
        print("データが取得できませんでした。")
        sys.exit(1)

    # --usage-year が指定されている場合、total_usage をその年の実績に置き換える
    if args.usage_year:
        for r in all_records:
            target_usage = 0
            for b in r.get("biblio", []):
                if str(b.get("year")) == str(args.usage_year):
                    try:
                        target_usage = int(b.get("yearToDateDownloads", 0))
                    except ValueError:
                        target_usage = 0
                    break
            r["total_usage"] = target_usage

    df_all = pd.DataFrame(all_records)
    print(f"\n総レコード数: {len(df_all)}")

    save_to_excel(df_all, all_records, args.output, args.usage_year)

    save_charts(df_all, args.output)

    # コンソールにも簡易サマリーを表示
    print("\n--- 年別合計サマリー ---")
    for journal_name in df_all["journal"].unique():
        df_j = df_all[df_all["journal"] == journal_name].dropna(subset=["year"])
        if df_j.empty:
            print(f"{journal_name}: データなし")
            continue
        yearly = df_j.groupby("year")[["total_usage", "citation_count", "reference_count", "ieice_ref_count"]].sum()
        print(f"\n{journal_name}:")
        print(yearly.to_string())


if __name__ == "__main__":
    main()
